import io
import re
from io import BytesIO
import pandas as pd
import streamlit as st
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Masterfile Filler", layout="wide")

# --- Header ---
st.markdown("<h1 style='text-align: center;'>🧩 Target Masterfile Filler</h1>", unsafe_allow_html=True)
st.markdown("<h4 style='text-align: center; font-style: italic;'>Innovating with AI Today ⏐ Leading Automation Tomorrow</h4>", unsafe_allow_html=True)

# ---------- Helpers ----------
def _norm_key(s: str) -> str:
    """Normalize header names: lowercase + remove non-alphanumerics."""
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def list_excel_sheets(uploaded_file) -> List[str]:
    if uploaded_file is None:
        return []
    try:
        xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
        return xls.sheet_names
    except Exception:
        return []

def read_raw(uploaded_file, sheet: Optional[str]) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame()
    name = uploaded_file.name.lower()
    try:
        if name.endswith((".csv", ".txt")):
            return pd.read_csv(uploaded_file)
        # Excel
        if sheet is not None:
            return pd.read_excel(uploaded_file, sheet_name=sheet, engine="openpyxl")
        # default first
        xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
        first = xls.sheet_names[0]
        return pd.read_excel(xls, sheet_name=first, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read RAW: {e}")
        return pd.DataFrame()

def read_two_col_mapping(uploaded_file) -> pd.DataFrame:
    """
    Expect exactly 2 columns in this order:
      1) Template header  (e.g., 'Template', 'Template Header', 'Header of Masterfile Template', 'Target', 'To')
      2) Raw header       (e.g., 'Raw', 'Raw Header', 'Header of Row Sheet', 'Source', 'From')
    Output DataFrame columns: ['raw_header','template_header'] (used by writer).
    """
    if uploaded_file is None:
        return pd.DataFrame()

    name = getattr(uploaded_file, "name", "").lower()
    try:
        if name.endswith((".csv", ".txt")):
            m = pd.read_csv(uploaded_file)
        else:
            m = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read mapping: {e}")
        return pd.DataFrame()
    if m.empty:
        return m

    cols_norm = { _norm_key(c): c for c in m.columns }
    template_keys = [_norm_key(x) for x in ["Template","Template Header","Header of Masterfile Template","Target","To"]]
    raw_keys      = [_norm_key(x) for x in ["Raw","Raw Header","Header of Row Sheet","Source","From"]]

    tpl_col_name = next((cols_norm[k] for k in template_keys if k in cols_norm), None)
    raw_col_name = next((cols_norm[k] for k in raw_keys if k in cols_norm), None)
    if not tpl_col_name or not raw_col_name:
        st.error("Mapping must contain two columns: (1) Template header, (2) Raw header.")
        return pd.DataFrame()

    out = m[[tpl_col_name, raw_col_name]].copy()
    out.columns = ["template_header", "raw_header"]
    out["template_header"] = out["template_header"].astype(str).str.strip()
    out["raw_header"] = out["raw_header"].astype(str).str.strip()
    out = out[(out["template_header"] != "") & (out["raw_header"] != "")]
    out = out.drop_duplicates(subset=["template_header"]).reset_index(drop=True)
    # writer expects ['raw_header','template_header']
    return out[["raw_header","template_header"]]

def build_header_index_first_sheet(ws, header_row: int = 1) -> Dict[str, int]:
    """Return normalized header -> column index for FIRST sheet (row 1)."""
    headers = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = _norm_key(v)
        if key and key not in headers:
            headers[key] = c
    return headers

def _highlight_duplicates(ws, header_map: Dict[str, int], header_labels: List[str], start_row: int = 3):
    """
    For each header in header_labels, locate the column and highlight duplicate cells (yellow)
    from start_row to the last non-empty row. Case-insensitive, ignores blanks.
    """
    dup_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for hdr in header_labels:
        key = _norm_key(hdr)
        col_idx = header_map.get(key)
        if not col_idx:
            continue  # header not found; skip

        counts = {}
        max_row = ws.max_row or start_row
        for r in range(start_row, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            k = s.upper()
            counts[k] = counts.get(k, 0) + 1

        if counts:
            for r in range(start_row, max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                if counts.get(s.upper(), 0) > 1:
                    ws.cell(row=r, column=col_idx).fill = dup_fill

def fill_first_sheet_by_headers(template_bytes: BytesIO, mapping_df: pd.DataFrame, raw_df: pd.DataFrame, template_filename: str) -> BytesIO:
    """
    Write values ONLY on the first sheet using headers in row 1 (normalized match).
    Start writing from row 3. After filling, highlight duplicates in 'Partner SKU' and 'Barcode'.
    Other sheets remain untouched (names, formatting, formulas, merges).
    """
    keep_vba = template_filename.lower().endswith(".xlsm")
    wb = load_workbook(filename=template_bytes, data_only=False, keep_vba=keep_vba)
    ws = wb.worksheets[0]  # FIRST sheet only

    # Header index
    tpl_header_to_col = build_header_index_first_sheet(ws, header_row=1)
    if not tpl_header_to_col:
        raise ValueError("No headers found in row 1 of the first sheet. Please ensure row 1 contains headers.")

    # Raw column lookup (literal lowercase)
    raw_norm = {c.strip().lower(): c for c in raw_df.columns}

    # Build mapping pairs
    pairs = []
    missing_raw, missing_tpl = [], []
    for _, r in mapping_df.iterrows():
        raw_hdr_lc = r["raw_header"].strip().lower()
        tpl_hdr_norm = _norm_key(r["template_header"])
        raw_col_name = raw_norm.get(raw_hdr_lc)
        col_idx = tpl_header_to_col.get(tpl_hdr_norm)
        if raw_col_name is None:
            missing_raw.append(r["raw_header"]); continue
        if col_idx is None:
            missing_tpl.append(r["template_header"]); continue
        pairs.append((raw_col_name, col_idx))

    if missing_tpl:
        st.warning("Template headers not found in row 1 (skipped): " + ", ".join(sorted(set(missing_tpl))))
    if missing_raw:
        st.warning("RAW columns missing (skipped): " + ", ".join(sorted(set(missing_raw))))

    # Write rows from row 3
    start_row = 3
    for i, (_, raw_row) in enumerate(raw_df.iterrows()):
        out_row = start_row + i
        for raw_col_name, col_idx in pairs:
            val = raw_row[raw_col_name]
            ws.cell(row=out_row, column=col_idx, value=("" if pd.isna(val) else val))

    # Highlight duplicates in specific columns
    _highlight_duplicates(ws, tpl_header_to_col, header_labels=["Partner SKU", "Barcode"], start_row=start_row)

    # Save to bytes
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------- SESSION ----------
if "template_bytes" not in st.session_state:
    st.session_state.template_bytes = None
if "template_name" not in st.session_state:
    st.session_state.template_name = None
if "raw_df" not in st.session_state:
    st.session_state.raw_df = pd.DataFrame()

# ---------- UI (tabs in requested order) ----------
tab1, tab2, tab3 = st.tabs([
    "1) Upload Masterfile Template",
    "2) Upload Raw Data",
    "3) Upload Mapping & Download"
])

# Tab 1: Template
with tab1:
    st.subheader("Upload Masterfile Template (XLSX or XLSM)")
    template_file = st.file_uploader("Masterfile (Excel .xlsx or .xlsm)", type=["xlsx","xlsm"], key="template_file")
    if template_file is not None:
        st.session_state.template_bytes = BytesIO(template_file.getbuffer())
        st.session_state.template_name = template_file.name
        try:
            xls = pd.ExcelFile(template_file, engine="openpyxl")
            first_sheet_name = xls.sheet_names[0]
            tpl_preview = pd.read_excel(xls, sheet_name=first_sheet_name, nrows=5, engine="openpyxl")
            st.info(f"Only the FIRST sheet will be modified: **{first_sheet_name}** (Row 3 onward).")
            st.dataframe(tpl_preview, use_container_width=True)
        except Exception as e:
            st.error(f"Failed to read template for preview: {e}")

# Tab 2: Raw
with tab2:
    st.subheader("Upload Raw Data (CSV/XLSX)")
    raw_file = st.file_uploader("Raw file", type=["csv","xlsx"], key="raw_file")
    raw_sheet = None
    if raw_file is not None and raw_file.name.lower().endswith(".xlsx"):
        try:
            sheets = pd.ExcelFile(raw_file, engine="openpyxl").sheet_names
            if sheets:
                raw_sheet = st.selectbox("Select RAW sheet", options=sheets, index=0)
        except Exception as e:
            st.error(f"Could not read RAW Excel: {e}")

    raw_df = read_raw(raw_file, raw_sheet)
    if not raw_df.empty:
        st.session_state.raw_df = raw_df
        st.success(f"Loaded RAW: {len(raw_df):,} rows × {len(raw_df.columns):,} columns.")
        st.dataframe(raw_df.head(50), use_container_width=True)

# Tab 3: Mapping + Process
with tab3:
    st.subheader("Upload 2-column mapping (Template, Raw) & Generate")
    st.markdown(
        "Columns required (in this order):  \n"
        "- **Template** (first column): header text in the template's FIRST sheet (row 1)  \n"
        "- **Raw** (second column): column name in your raw file  \n\n"
        "The app will copy **Raw → Template** and then highlight duplicates in **Partner SKU** and **Barcode**."
    )
    mapping_file = st.file_uploader("Mapping file (XLSX/CSV)", type=["xlsx","csv"], key="mapping_file")

    # Mapping template download
    def mapping_template_bytes():
        df = pd.DataFrame({
            "Template": ["SKU","Title","Description","Price","Quantity","Category"],
            "Raw":      ["sku","name","description","price","qty","category"],
        })
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name="mapping")
        out.seek(0); return out

    st.download_button("⬇️ Download mapping_template.xlsx",
                       data=mapping_template_bytes().getvalue(),
                       file_name="mapping_template.xlsx")

    can_process = (
        st.session_state.template_bytes is not None
        and st.session_state.template_name is not None
        and not st.session_state.raw_df.empty
        and mapping_file is not None
    )
    if not can_process:
        st.info("Please upload Template (Tab 1), Raw (Tab 2), and Mapping (Tab 3) to enable processing.")

    if st.button("⚙️ Process the Data", type="primary", disabled=not can_process):
        try:
            # Read mapping directly from UploadedFile (avoid BytesIO -> no .name error)
            mapping_df = read_two_col_mapping(mapping_file)
            if mapping_df.empty:
                st.error("Mapping is empty or invalid.")
            else:
                out_bytes = fill_first_sheet_by_headers(
                    template_bytes=BytesIO(st.session_state.template_bytes.getbuffer()),
                    mapping_df=mapping_df,
                    raw_df=st.session_state.raw_df,
                    template_filename=st.session_state.template_name
                )
                st.success("Done! Your updated masterfile is ready. Duplicate Partner SKU / Barcode cells are highlighted.")
                st.download_button(
                    label="⬇️ Download Updated Masterfile (Excel)",
                    data=out_bytes.getvalue(),
                    file_name="filled_masterfile.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.exception(e)
